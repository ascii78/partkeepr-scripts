#!/usr/bin/env python3

# Example partkeepr csv to libreoffice calc
# -----------------------------------------
#
# NOTE: not tested on excel
#

import csv
# for json pretty printing
from pprint import pprint as pp

# http://docs.python-requests.org/en/master/
import requests

# https://openpyxl.readthedocs.io/
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# set this to partkeepr settings
HOST = "http://localhost/api"
USER = "admin"
PASS = "admin"


dest_filename = 'export.xlsx'

# First workbook
wb = Workbook()

# First sheet
ws1 = wb.active
ws1.title = "Import"

column_headers = [
    "Name",
    "Description",
    "Location",
    "Category",
    "Category_Name",
    "Stock",
    "Manufacturer",
    "MPN",
    "Distributor",
    "SKU",
    "Price",
    "Order",
    "Link"
]

ws1.append(column_headers)

# Second sheet is categories, example JSON request
ws2 = wb.create_sheet(title="Categories")

# Set this to enable csv exporting
headers = {'Accept': 'text/comma-separated-values'}
auth = (USER, PASS)

# Export settings, note the columns entry, it's actually a string 
# but it looks like a dict
params = (
    ('_dc', '1506284694809'),
    ('itemsPerPage', '9999999'),
    ('columns', '["level", "name", "categoryPath"]'),
)

# issue http GET
r = requests.get('%s/part_categories' % HOST, params=params, headers=headers, auth=auth)

# Optionally use r.status_code, r.text, pp(r.json())
# we just want to parse csv
reader = csv.reader(r.text.splitlines())
for row in reader:
    # pp(row)
    # we only want end nodes, my categories are only 3 deep, also want the header
    ws2.append((row[2], row[1]))

# set data validation, also known as a dropdown menu for the categories column,
# use complete path
dv = DataValidation(type="list",
                    formula1="=Categories!$A$2:$A$1048576")
dv.ranges.add('D2:D1048576')
ws1.add_data_validation(dv)

# setup some optional formulas
for i in range(2, 1000):
    # setup clickable url if SKU entry exists, search for url in distributors list and combine the two
    ws1["M{}".format(i)] = '=IF(J{}="","",HYPERLINK(VLOOKUP(I{}, Distributor!$A$2:$B$45, 2)&J{},I{})'.format(i, i, i, i)
    # return name of a complete category path, probably easier for importing later
    ws1["E{}".format(i)] = '=IF(D{}="","",VLOOKUP(D{}, Categories!$A$2:$B$45, 2, 0))'.format(i, i)

# unfortunately can't really auto size these things
ws1.column_dimensions["B"].width = 50
ws1.column_dimensions["C"].width = 35
ws1.column_dimensions["D"].width = 100
ws1.column_dimensions["E"].width = 40
ws1.column_dimensions["G"].width = 35
ws1.column_dimensions["I"].width = 35

# better make this a function
def get_csv_export(name, endpoint, columns):
    print(name, endpoint, columns)
    ws = wb.create_sheet(title=name)

    params = (
        ('_dc', '1506284694809'),
        ('itemsPerPage', '9999999'),
        ('columns', columns),
    )
    r = requests.get('%s/%s' % (HOST, endpoint), params=params, headers=headers, auth=auth)
    reader = csv.reader(r.text.splitlines())
    for row in reader:
        if name == "Distributor":
            row[1] = str(row[1]).replace("%s", "")
        ws.append(row)
    dv = DataValidation(type="list",
                        formula1="={}!$A$2:$A$1048576".format(name))
    # get column name based on header using ascii
    col = chr(column_headers.index(name) + 97).upper()
    # apply datavalidation to first sheet
    dv.ranges.add('{}2:{}1048576'.format(col, col))
    # save data validation to first sheet
    ws1.add_data_validation(dv)


get_csv_export("Location", "storage_locations", '["name"]')
get_csv_export("Distributor", "distributors", '["name", "skuurl"]')
get_csv_export("Manufacturer", "manufacturers", '["name"]')

wb.save(filename=dest_filename)
